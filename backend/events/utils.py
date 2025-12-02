import openpyxl
from io import BytesIO
from django.conf import settings
import os

def generate_event_report(data, template_name="template_relatorio.xlsx"):
    file_path = os.path.join(settings.BASE_DIR, 'events', 'templates', template_name)
    
    wb = openpyxl.load_workbook(file_path)
    
    participants = data.get('participants', [])

    if 'Dados Gerais' in wb.sheetnames:
        wd = wb['Dados Gerais']
        start_row = 8
        
        wd['C2'] = data.get('event_name')   
        wd['C3'] = data.get('period')
        wd['C4'] = data.get('location')
        
        for i, p in enumerate(participants):
            row = start_row + i
            wd.cell(row=row, column=2, value=p.get('name'))
            wd.cell(row=row, column=3, value=p.get('birth_date'))
            wd.cell(row=row, column=4, value=p.get('cpf'))
            wd.cell(row=row, column=7, value=p.get('role'))
            wd.cell(row=row, column=8, value=p.get('email'))
            wd.cell(row=row, column=9, value=p.get('phone'))

    if 'Passagens' in wb.sheetnames:
        wp = wb['Passagens']
        start_row = 7
        
        for i, p in enumerate(participants):
            row = start_row + i
            wp.cell(row=row, column=5, value=p.get('departure_date'))
            wp.cell(row=row, column=6, value=p.get('origin'))
            wp.cell(row=row, column=10, value=p.get('return_date'))
            
    if 'Hospedagem' in wb.sheetnames:
        wh = wb['Hospedagem']
        start_row = 7
        
        for i, p in enumerate(participants):
            row = start_row + i
            wh.cell(row=row, column=5, value=p.get('room_type'))

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer